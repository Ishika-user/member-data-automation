"""
Extraction module for the Automated Member Data Migration and QA System.

Responsible for safely reading the source workbook, identifying the correct
sheet, mapping loosely-formatted headers to standardized field names, and
returning a clean DataFrame plus extraction metadata. This module never
modifies the original file and never logs full member records.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pandas as pd

from src.config import HEADER_MATCH_RULES, STANDARD_SOURCE_COLUMNS


# ---------------------------------------------------------------------------
# Custom errors
# ---------------------------------------------------------------------------
class ExtractionError(Exception):
    """Base class for extraction-related errors."""


class FileNotFoundExtractionError(ExtractionError):
    pass


class UnsupportedFileTypeError(ExtractionError):
    pass


class WorkbookReadError(ExtractionError):
    pass


class MissingColumnsError(ExtractionError):
    pass


class AmbiguousHeadingError(ExtractionError):
    pass


class EmptyWorkbookError(ExtractionError):
    pass


# ---------------------------------------------------------------------------
# Metadata container
# ---------------------------------------------------------------------------
@dataclass
class ExtractionMetadata:
    selected_sheet_label: str
    selected_sheet_index: int
    input_row_count: int
    column_mapping: Dict[str, str] = field(default_factory=dict)
    sheet_names_generic: List[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _normalize_header(raw: object) -> str:
    """Normalize a header value for tolerant matching."""
    if raw is None:
        return ""
    text = str(raw)
    text = text.strip()
    text = re.sub(r"\s+", " ", text)
    text = text.lower()
    # Ignore harmless punctuation differences (periods, #, apostrophes)
    text = re.sub(r"[.#']", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _match_headers(headers: List[object]) -> Dict[str, int]:
    """
    Attempt to map normalized headers to standardized field names.
    Returns a dict of standardized_name -> column_index (0-based).
    Only the first confident match per field is used.
    """
    normalized = [_normalize_header(h) for h in headers]
    mapping: Dict[str, int] = {}

    for std_name, fragments in HEADER_MATCH_RULES.items():
        candidates = []
        for idx, norm in enumerate(normalized):
            if not norm:
                continue
            for fragment in fragments:
                if fragment in norm:
                    candidates.append(idx)
                    break
        if len(candidates) == 1:
            mapping[std_name] = candidates[0]
        elif len(candidates) > 1:
            # Prefer the shortest header text (most exact match) among candidates
            candidates.sort(key=lambda i: len(normalized[i]))
            mapping[std_name] = candidates[0]
    return mapping


def _sheet_confidence(headers: List[object]) -> int:
    """Count how many required standardized columns a sheet's headers satisfy."""
    return len(_match_headers(headers))


# ---------------------------------------------------------------------------
# Main extraction function
# ---------------------------------------------------------------------------
def extract_members(input_path: Path, logger=None) -> "tuple[pd.DataFrame, ExtractionMetadata]":
    """
    Extract the four required member fields from the source workbook.

    Returns a tuple of (DataFrame, ExtractionMetadata). The DataFrame
    contains columns: original_member_no, member_name, nepali_name, gender,
    source_row_number.
    """
    input_path = Path(input_path)

    if not input_path.exists():
        raise FileNotFoundExtractionError(f"Input file does not exist: {input_path.name}")

    if input_path.suffix.lower() not in (".xlsx", ".xls"):
        raise UnsupportedFileTypeError(
            f"Unsupported file type '{input_path.suffix}'. Expected .xlsx or .xls."
        )

    try:
        workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced as a clear custom error
        raise WorkbookReadError(
            f"Unable to read workbook (corrupted, password-protected, or invalid format): {exc}"
        ) from exc

    sheet_names = workbook.sheetnames
    if not sheet_names:
        raise EmptyWorkbookError("Workbook contains no sheets.")

    generic_labels = [f"Sheet_{i + 1}" for i in range(len(sheet_names))]

    # Evaluate each sheet's header row for column-matching confidence.
    best_index: Optional[int] = None
    best_score = -1
    best_headers: List[object] = []

    for idx, name in enumerate(sheet_names):
        ws = workbook[name]
        try:
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            continue
        score = _sheet_confidence(list(first_row))
        if logger:
            logger.info(
                "Sheet evaluation: label=%s required_columns_matched=%d",
                generic_labels[idx],
                score,
            )
        if score > best_score:
            best_score = score
            best_index = idx
            best_headers = list(first_row)

    if best_index is None or best_score < len(STANDARD_SOURCE_COLUMNS):
        raise MissingColumnsError(
            "Could not confidently identify all four required columns "
            "(original_member_no, member_name, nepali_name, gender) in any sheet."
        )

    mapping = _match_headers(best_headers)
    if len(mapping) < len(STANDARD_SOURCE_COLUMNS):
        missing = set(STANDARD_SOURCE_COLUMNS) - set(mapping.keys())
        raise MissingColumnsError(
            f"Missing confident mapping for required field(s): {sorted(missing)}"
        )

    selected_sheet_name = sheet_names[best_index]
    ws = workbook[selected_sheet_name]

    records = []
    source_row_number = 1  # header is row 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        source_row_number += 1
        if row is None or all(v is None for v in row):
            continue
        record = {
            "original_member_no": row[mapping["original_member_no"]]
            if mapping["original_member_no"] < len(row)
            else None,
            "member_name": row[mapping["member_name"]] if mapping["member_name"] < len(row) else None,
            "nepali_name": row[mapping["nepali_name"]] if mapping["nepali_name"] < len(row) else None,
            "gender": row[mapping["gender"]] if mapping["gender"] < len(row) else None,
            "source_row_number": source_row_number,
        }
        # Skip fully blank data rows (no member number and no name)
        if record["original_member_no"] is None and record["member_name"] is None:
            continue
        records.append(record)

    workbook.close()

    df = pd.DataFrame(records, columns=STANDARD_SOURCE_COLUMNS + ["source_row_number"])

    metadata = ExtractionMetadata(
        selected_sheet_label=generic_labels[best_index],
        selected_sheet_index=best_index,
        input_row_count=len(df),
        column_mapping={k: generic_labels[best_index] + f"[col {v}]" for k, v in mapping.items()},
        sheet_names_generic=generic_labels,
    )

    return df, metadata
