"""
Report generation module for the Automated Member Data Migration and QA System.

Builds professionally formatted Excel outputs (cleaned members, manual
review, and the multi-sheet QA report) using openpyxl styling.
"""

from __future__ import annotations

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def _style_header(ws: Worksheet, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autofit_columns(ws: Worksheet, df_or_rows, num_columns: int) -> None:
    widths = [12] * num_columns
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                length = len(str(cell.value))
                col_idx = cell.column - 1
                if col_idx < num_columns:
                    widths[col_idx] = max(widths[col_idx], min(length + 2, 45))
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_dataframe(ws: Worksheet, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    _style_header(ws, len(df.columns))
    _autofit_columns(ws, df, len(df.columns))


def write_simple_workbook(df: pd.DataFrame, output_path: Path, sheet_title: str = "Data") -> None:
    """Write a single-sheet formatted Excel workbook from a DataFrame."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    _write_dataframe(ws, df)
    wb.save(output_path)


def write_qa_report(
    output_path: Path,
    summary_rows: List[dict],
    missing_values_rows: List[dict],
    duplicate_summary_rows: List[dict],
    duplicate_details_rows: List[dict],
    gender_distribution_rows: List[dict],
    validation_results: List[dict],
    run_metadata_rows: List[dict],
) -> None:
    """Write the full multi-worksheet QA report with professional formatting."""
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    summary_df = pd.DataFrame(summary_rows)
    _write_dataframe(ws, summary_df)

    # Missing Values
    ws = wb.create_sheet("Missing Values")
    df = pd.DataFrame(missing_values_rows)
    _write_dataframe(ws, df)

    # Duplicate Summary
    ws = wb.create_sheet("Duplicate Summary")
    df = pd.DataFrame(duplicate_summary_rows)
    _write_dataframe(ws, df)

    # Duplicate Details
    ws = wb.create_sheet("Duplicate Details")
    df = pd.DataFrame(duplicate_details_rows) if duplicate_details_rows else pd.DataFrame(
        columns=["duplicate_type", "original_member_no", "source_row_number", "status", "reason"]
    )
    _write_dataframe(ws, df)

    # Gender Distribution
    ws = wb.create_sheet("Gender Distribution")
    df = pd.DataFrame(gender_distribution_rows)
    _write_dataframe(ws, df)

    # Validation Results
    ws = wb.create_sheet("Validation Results")
    df = pd.DataFrame(validation_results)
    df = df.rename(
        columns={
            "rule": "Validation Rule",
            "expected": "Expected Result",
            "actual": "Actual Result",
            "passed": "Passed",
            "explanation": "Explanation",
        }
    )
    _write_dataframe(ws, df)
    # Highlight passed/failed/warning cells in the "Passed" column
    passed_col_idx = list(df.columns).index("Passed") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=passed_col_idx)
        if cell.value == "Passed":
            cell.fill = PASS_FILL
        elif cell.value == "Failed":
            cell.fill = FAIL_FILL
        elif cell.value == "Warning":
            cell.fill = WARN_FILL

    # Run Metadata
    ws = wb.create_sheet("Run Metadata")
    df = pd.DataFrame(run_metadata_rows)
    _write_dataframe(ws, df)

    wb.save(output_path)
